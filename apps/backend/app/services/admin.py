from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from apps.backend.app.schemas.admin import AdminOverview, AdminUserRow
from database.schema.models import Goal, Transaction, TransactionType, User


def build_admin_overview(db: Session) -> AdminOverview:
    users = list(db.scalars(select(User).order_by(User.created_at.desc())))
    transactions = list(db.scalars(select(Transaction)))
    goals = list(db.scalars(select(Goal)))

    transaction_count_by_user: dict[int, int] = {}
    spending_by_user: dict[int, float] = {}
    income_by_user: dict[int, float] = {}
    for item in transactions:
        transaction_count_by_user[item.user_id] = transaction_count_by_user.get(item.user_id, 0) + 1
        if item.transaction_type == TransactionType.INCOME:
            income_by_user[item.user_id] = income_by_user.get(item.user_id, 0.0) + float(item.amount)
        else:
            spending_by_user[item.user_id] = spending_by_user.get(item.user_id, 0.0) + float(item.amount)

    goal_count_by_user: dict[int, int] = {}
    target_by_user: dict[int, float] = {}
    for goal in goals:
        goal_count_by_user[goal.user_id] = goal_count_by_user.get(goal.user_id, 0) + 1
        target_by_user[goal.user_id] = target_by_user.get(goal.user_id, 0.0) + float(goal.target_amount)

    user_rows = [
        AdminUserRow(
            id=user.id,
            full_name=user.full_name,
            email=user.email,
            is_active=user.is_active,
            created_at=user.created_at,
            transaction_count=transaction_count_by_user.get(user.id, 0),
            total_spent=round(spending_by_user.get(user.id, 0.0), 2),
            total_income=round(income_by_user.get(user.id, 0.0), 2),
            goal_count=goal_count_by_user.get(user.id, 0),
            total_goal_target=round(target_by_user.get(user.id, 0.0), 2),
        )
        for user in users
    ]

    return AdminOverview(
        total_users=len(users),
        total_transactions=len(transactions),
        total_goals=len(goals),
        total_spent=round(
            sum(float(item.amount) for item in transactions if item.transaction_type == TransactionType.EXPENSE),
            2,
        ),
        total_income=round(
            sum(float(item.amount) for item in transactions if item.transaction_type == TransactionType.INCOME),
            2,
        ),
        users=user_rows,
    )


def build_admin_export_workbook(db: Session) -> bytes:
    workbook = Workbook()

    users_sheet = workbook.active
    users_sheet.title = "Users"
    _write_sheet(
        users_sheet,
        [
            "ID",
            "Full Name",
            "Email",
            "Active",
            "Created At",
        ],
        [
            [user.id, user.full_name, user.email, user.is_active, user.created_at.isoformat()]
            for user in db.scalars(select(User).order_by(User.created_at.desc()))
        ],
    )

    transactions_sheet = workbook.create_sheet("Transactions")
    _write_sheet(
        transactions_sheet,
        [
            "ID",
            "User ID",
            "Title",
            "Category",
            "Transaction Type",
            "Amount",
            "Transaction Date",
            "Notes",
            "Created At",
        ],
        [
            [
                item.id,
                item.user_id,
                item.title,
                item.category,
                item.transaction_type.value,
                float(item.amount),
                item.transaction_date.isoformat(),
                item.notes or "",
                item.created_at.isoformat(),
            ]
            for item in db.scalars(select(Transaction).order_by(Transaction.transaction_date.desc()))
        ],
    )

    goals_sheet = workbook.create_sheet("Goals")
    _write_sheet(
        goals_sheet,
        [
            "ID",
            "User ID",
            "Title",
            "Description",
            "Target Amount",
            "Current Amount",
            "Status",
            "Created At",
        ],
        [
            [
                goal.id,
                goal.user_id,
                goal.title,
                goal.description or "",
                float(goal.target_amount),
                float(goal.current_amount),
                str(goal.status.value),
                goal.created_at.isoformat(),
            ]
            for goal in db.scalars(select(Goal).order_by(Goal.created_at.desc()))
        ],
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_sheet(sheet: Any, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(row)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
